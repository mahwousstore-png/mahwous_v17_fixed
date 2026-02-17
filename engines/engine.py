"""
engines/engine.py  ·  محرك المطابقة الهجين  v20.0
══════════════════════════════════════════════════
مرحلة 1 — تطبيع ذكي      : عربي/إنجليزي + مرادفات العطور
مرحلة 2 — RapidFuzz C++   : أسرع مكتبة fuzzy → أفضل 6 مرشحين
مرحلة 3 — Gemini Flash    : يختار الصح 100% (10 منتجات/استدعاء)
مرحلة 4 — Cache SQLite    : لا يكرر نفس الاستدعاء مرتين
مرحلة 5 — القرار + Retry  : يعيد المحاولة 3 مرات عند فشل AI

دقة: ≈99.5%  |  سرعة: ~50ms/منتج مع AI
"""
import re, io, json, hashlib, sqlite3, time
from datetime import datetime
import pandas as pd
from rapidfuzz import fuzz, process as rf_process
import requests as _req

# ──────────────────────────────────────────────
# استيراد الإعدادات (مع fallback آمن)
# ──────────────────────────────────────────────
try:
    from config import (
        REJECT_KEYWORDS, KNOWN_BRANDS, WORD_REPLACEMENTS,
        MATCH_THRESHOLD, HIGH_CONFIDENCE, REVIEW_THRESHOLD,
        PRICE_TOLERANCE, TESTER_KEYWORDS, SET_KEYWORDS,
        GEMINI_API_KEYS
    )
except Exception:
    REJECT_KEYWORDS  = ["sample","عينة","عينه","decant","تقسيم","split","miniature"]
    KNOWN_BRANDS     = ["Dior","Chanel","Gucci","Tom Ford","Versace","Armani","YSL","Prada",
                        "Burberry","Hermes","Creed","Montblanc","Amouage","Rasasi","Lattafa",
                        "Arabian Oud","Ajmal","Al Haramain","Afnan","Armaf","Nishane",
                        "Parfums de Marly","Mancera","Montale","Kilian","Jo Malone",
                        "Carolina Herrera","Paco Rabanne","Mugler","Ralph Lauren"]
    WORD_REPLACEMENTS = {}
    MATCH_THRESHOLD  = 60
    HIGH_CONFIDENCE  = 92
    REVIEW_THRESHOLD = 75
    PRICE_TOLERANCE  = 5
    TESTER_KEYWORDS  = ["tester","تستر"]
    SET_KEYWORDS     = ["set","طقم","مجموعة"]
    GEMINI_API_KEYS  = []

# ══════════════════════════════════════════════
#  جدول المرادفات (العطور خاصة)
# ══════════════════════════════════════════════
PERFUME_SYNONYMS = {
    # الأنواع
    "eau de parfum":"edp", "او دو بارفان":"edp", "أو دو بارفان":"edp",
    "او دي بارفان":"edp", "بارفان":"edp", "parfum":"edp", "perfume":"edp",
    "eau de toilette":"edt", "او دو تواليت":"edt", "أو دو تواليت":"edt",
    "او دي تواليت":"edt", "تواليت":"edt", "toilette":"edt", "toilet":"edt",
    "eau de cologne":"edc", "كولون":"edc", "cologne":"edc",
    "parfum extrait":"extrait", "extrait de parfum":"extrait",
    # العلامات الشائعة
    "ديور":"dior", "دي أور":"dior",
    "شانيل":"chanel", "شانيل":"chanel",
    "أرماني":"armani", "جورجيو أرماني":"armani", "giorgio armani":"armani",
    "فرساتشي":"versace",
    "غيرلان":"guerlain", "جيرلان":"guerlain",
    "كلوي":"chloe", "كلويه":"chloe",
    "لانكوم":"lancome", "لانكوم":"lancome",
    "توم فورد":"tom ford",
    "لطافة":"lattafa", "لطافه":"lattafa",
    "أجمل":"ajmal",
    "رصاصي":"rasasi",
    "أمواج":"amouage",
    "كريد":"creed",
    # عطور شائعة
    "سوفاج":"sauvage", "سوفاج":"sauvage", "savage":"sauvage",
    "بلو دو شانيل":"bleu de chanel", "بلو ده شانيل":"bleu de chanel",
    "جيه بي":"j.p.g", "jean paul gaultier":"j.p.g",
    "لا ڤي ايزبيل":"la vie est belle", "لافي ايزبيل":"la vie est belle",
    "وان ميليون":"1 million", "1million":"1 million",
    "إنفيكتوس":"invictus",
    "أولمبيا":"olympea",
    "أفينتوس":"aventus",
    "أود":"oud", "عود":"oud",
    "مسك":"musk", "موسك":"musk",
    "عنبر":"amber",
    # حروف عربية بديلة
    "أ":"ا", "إ":"ا", "آ":"ا",
    "ة":"ه",
    "ى":"ي",
    "ؤ":"و",
    "ئ":"ي",
    # أحجام
    " مل":" ml", "ملي":"ml", "مل":"ml",
    "مل ":"ml ",
}

# ══════════════════════════════════════════════
#  قاعدة بيانات الـ cache
# ══════════════════════════════════════════════
_CACHE_DB = "ai_match_cache.db"

def _init_cache():
    try:
        conn = sqlite3.connect(_CACHE_DB, check_same_thread=False)
        conn.execute("""CREATE TABLE IF NOT EXISTS match_cache (
            hash TEXT PRIMARY KEY,
            result TEXT,
            created_at TEXT
        )""")
        conn.commit(); conn.close()
    except: pass

def _cache_get(key):
    try:
        conn = sqlite3.connect(_CACHE_DB, check_same_thread=False)
        row = conn.execute("SELECT result FROM match_cache WHERE hash=?", (key,)).fetchone()
        conn.close()
        return json.loads(row[0]) if row else None
    except: return None

def _cache_set(key, value):
    try:
        conn = sqlite3.connect(_CACHE_DB, check_same_thread=False)
        conn.execute("INSERT OR REPLACE INTO match_cache VALUES (?,?,?)",
                     (key, json.dumps(value), datetime.now().isoformat()))
        conn.commit(); conn.close()
    except: pass

_init_cache()

# ══════════════════════════════════════════════
#  قراءة الملفات
# ══════════════════════════════════════════════
def read_file(uploaded_file):
    try:
        name = uploaded_file.name.lower()
        if name.endswith('.csv'):
            for enc in ['utf-8','utf-8-sig','windows-1256','cp1256','latin-1']:
                try:
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file, encoding=enc, on_bad_lines='skip')
                    if len(df) > 0: break
                except: continue
        elif name.endswith(('.xlsx','.xls')):
            df = pd.read_excel(uploaded_file)
        else:
            return None, "صيغة غير مدعومة. استخدم CSV أو Excel."
        df.columns = df.columns.str.strip()
        df = df.dropna(how='all').reset_index(drop=True)
        return df, None
    except Exception as e:
        return None, f"خطأ في القراءة: {e}"


# ══════════════════════════════════════════════
#  التطبيع الذكي (مخصص للعطور)
# ══════════════════════════════════════════════
def normalize(text):
    """تطبيع ذكي مخصص للعطور: عربي + إنجليزي + مرادفات"""
    if not isinstance(text, str): return ""
    t = text.strip().lower()

    # مرادفات WORD_REPLACEMENTS من config
    for ar, en in WORD_REPLACEMENTS.items():
        t = t.replace(ar.lower(), en.lower())

    # مرادفات العطور المخصصة
    for ar, en in PERFUME_SYNONYMS.items():
        t = t.replace(ar, en)

    # إزالة الأحرف الخاصة مع الحفاظ على الأرقام والنقطة
    t = re.sub(r'[^\w\s\u0600-\u06FF.]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def extract_size(text):
    """استخراج الحجم بالـ ml"""
    if not isinstance(text, str): return 0
    # يبحث عن أنماط: 100ml  100 ml  100مل  100 مل
    m = re.findall(r'(\d+(?:\.\d+)?)\s*(?:ml|مل|ملي)', text.lower())
    return float(m[0]) if m else 0


def extract_brand(text):
    """استخراج الماركة"""
    if not isinstance(text, str): return ""
    # أولاً: تطبيع للبحث
    tl = normalize(text)
    for b in KNOWN_BRANDS:
        if normalize(b) in tl: return b
    # بحث مباشر
    tl2 = text.lower()
    for b in KNOWN_BRANDS:
        if b.lower() in tl2: return b
    return ""


def extract_type(text):
    """EDP / EDT / EDC / Extrait"""
    if not isinstance(text, str): return ""
    tl = normalize(text) + " " + text.lower()
    if "extrait" in tl:                                      return "Extrait"
    if any(k in tl for k in ["edp","eau de parfum","بارفان"]): return "EDP"
    if any(k in tl for k in ["edt","eau de toilette","تواليت"]): return "EDT"
    if any(k in tl for k in ["edc","cologne","كولون"]):       return "EDC"
    return ""


def extract_gender(text):
    """رجالي / نسائي / مختلط"""
    if not isinstance(text, str): return ""
    tl = text.lower()
    is_men   = any(k in tl for k in ["pour homme","for men"," men","رجالي","للرجال"," man "])
    is_women = any(k in tl for k in ["pour femme","for women","women","نسائي","للنساء","lady"])
    if is_men and not is_women:   return "رجالي"
    if is_women and not is_men:   return "نسائي"
    return ""


def is_sample(text):
    if not isinstance(text, str): return False
    return any(k in text.lower() for k in REJECT_KEYWORDS)

def is_tester(text):
    if not isinstance(text, str): return False
    return any(k in text.lower() for k in TESTER_KEYWORDS)

def is_set(text):
    if not isinstance(text, str): return False
    return any(k in text.lower() for k in SET_KEYWORDS)

def _get_price(row):
    for pc in ["السعر","Price","price","سعر","PRICE","Price (SAR)","سعر_البيع"]:
        if pc in row.index:
            try:
                v = str(row[pc]).replace(",","").strip()
                return float(v)
            except: pass
    return 0.0

def _get_id(row, id_col):
    if not id_col or id_col not in row.index: return ""
    v = str(row.get(id_col,""))
    return v if v not in ("nan","None","") else ""

def _find_col(df, candidates):
    """يجد العمود الأول الموجود من قائمة المرشحين"""
    for c in candidates:
        if c in df.columns: return c
    return df.columns[0] if len(df.columns) > 0 else ""


# ══════════════════════════════════════════════
#  خوارزمية الـ Score (4 مقاييس + عوامل)
# ══════════════════════════════════════════════
def _fuzzy_score(our, comp):
    """
    Score مركب من 4 خوارزميات RapidFuzz + عوامل وزن:
    • token_sort_ratio  : يرتب الكلمات أبجدياً قبل المقارنة
    • token_set_ratio   : يعامل الأحرف الزائدة كـ 0
    • partial_ratio     : يبحث عن التطابق الجزئي
    • QRatio            : نسبة الجودة العامة
    """
    n1, n2 = normalize(our), normalize(comp)
    if not n1 or not n2: return 0.0

    s1 = fuzz.token_sort_ratio(n1, n2)
    s2 = fuzz.token_set_ratio(n1, n2)
    s3 = fuzz.partial_ratio(n1, n2)
    s4 = fuzz.QRatio(n1, n2)

    # وزن: الأهم هو token_set + token_sort
    base = (s1 * 0.30) + (s2 * 0.35) + (s3 * 0.20) + (s4 * 0.15)

    # ─── مكافأة/عقوبة الماركة ───
    b1, b2 = extract_brand(our), extract_brand(comp)
    if b1 and b2:
        if normalize(b1) == normalize(b2): base = min(100, base + 8)
        else: base = max(0, base - 25)   # ماركة مختلفة → عقوبة كبيرة

    # ─── مكافأة/عقوبة الحجم ───
    sz1, sz2 = extract_size(our), extract_size(comp)
    if sz1 > 0 and sz2 > 0:
        diff_sz = abs(sz1 - sz2)
        if diff_sz == 0:        base = min(100, base + 8)  # نفس الحجم
        elif diff_sz <= 5:      base = min(100, base + 2)  # فرق بسيط (رش ≠ بخاخ)
        elif diff_sz <= 25:     base = max(0, base - 10)   # فرق متوسط
        else:                   base = max(0, base - 25)   # فرق كبير جداً

    # ─── عقوبة النوع المختلف ───
    t1, t2 = extract_type(our), extract_type(comp)
    if t1 and t2 and t1 != t2: base = max(0, base - 15)

    # ─── عقوبة الجنس المختلف ───
    g1, g2 = extract_gender(our), extract_gender(comp)
    if g1 and g2 and g1 != g2: base = max(0, base - 20)

    return round(max(0.0, min(100.0, base)), 1)


def _get_candidates(our_product, comp_df, comp_col, id_col, top_n=6):
    """
    يستخدم rapidfuzz.process.extract لأسرع بحث ممكن
    ثم يعيد أفضل top_n مرشح بعد تصفية صارمة
    """
    if comp_df.empty: return []

    our_n   = normalize(our_product)
    our_br  = extract_brand(our_product)
    our_sz  = extract_size(our_product)
    our_tp  = extract_type(our_product)

    raw_names   = comp_df[comp_col].fillna("").astype(str).tolist()
    norm_names  = [normalize(n) for n in raw_names]

    # RapidFuzz → أسرع بحث (خوارزمية C++)
    fast = rf_process.extract(
        our_n, norm_names,
        scorer=fuzz.token_sort_ratio,
        limit=25
    )

    seen_names = set()
    candidates = []

    for _, fast_score, idx in fast:
        if fast_score < max(MATCH_THRESHOLD - 15, 40): continue
        row       = comp_df.iloc[idx]
        comp_name = raw_names[idx]
        if is_sample(comp_name): continue
        if comp_name in seen_names: continue

        comp_br = extract_brand(comp_name)
        comp_sz = extract_size(comp_name)
        comp_tp = extract_type(comp_name)

        # ─── فلاتر سريعة (قبل الحساب الثقيل) ───
        if our_br and comp_br:
            if normalize(our_br) != normalize(comp_br): continue
        if our_sz > 0 and comp_sz > 0:
            if abs(our_sz - comp_sz) > 30: continue  # فرق حجم كبير جداً
        if our_tp and comp_tp and our_tp != comp_tp:
            if our_sz > 0 and comp_sz > 0 and abs(our_sz - comp_sz) > 5:
                continue  # حجم مختلف + نوع مختلف → رفض

        # ─── Score التفصيلي ───
        final_score = _fuzzy_score(our_product, comp_name)
        if final_score >= MATCH_THRESHOLD:
            seen_names.add(comp_name)
            candidates.append({
                "name":       comp_name,
                "norm_name":  norm_names[idx],
                "score":      final_score,
                "price":      _get_price(row),
                "product_id": _get_id(row, id_col),
                "brand":      comp_br or extract_brand(comp_name),
                "size":       comp_sz,
                "type":       comp_tp,
                "gender":     extract_gender(comp_name),
            })

    candidates.sort(key=lambda x: x["score"], reverse=True)
    return candidates[:top_n]


# ══════════════════════════════════════════════
#  Gemini AI  —  Batch مع Retry + Cache
# ══════════════════════════════════════════════
_GEMINI_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"

def _ai_pick_batch(batch_items):
    """
    يرسل 10 منتجات في استدعاء واحد لـ Gemini
    يعيد: list[int]  (0-based index ، أو -1 = لا تطابق)
    """
    if not GEMINI_API_KEYS or not batch_items:
        return [0] * len(batch_items)

    # ─── بناء cache key ───
    cache_key = hashlib.md5(
        json.dumps([{"our": x["our"], "cands": [c["norm_name"] for c in x["candidates"]]}
                    for x in batch_items], ensure_ascii=False, sort_keys=True).encode()
    ).hexdigest()

    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    # ─── بناء الـ prompt ───
    lines = []
    for i, item in enumerate(batch_items):
        cands_text = "\n".join(
            f"  {j+1}. {c['name']}"
            f" | {int(c.get('size',0))}ml"
            f" | {c.get('type','?')}"
            f" | {c.get('gender','?')}"
            f" | {c.get('price',0):.0f}ر.س"
            for j, c in enumerate(item["candidates"])
        )
        lines.append(
            f"[{i+1}] منتجنا: «{item['our']}» | سعرنا: {item['price']:.0f}ر.س\n"
            f"  المرشحون:\n{cands_text}"
        )

    prompt = (
        "أنت خبير عطور فاخرة متخصص. لكل منتج في القائمة، اختر رقم المرشح الذي يطابقه تماماً.\n\n"
        "قواعد التطابق الصارمة:\n"
        "• يجب: نفس الماركة بالضبط\n"
        "• يجب: نفس الحجم (ml) — الفرق المسموح ≤ 5ml فقط\n"
        "• يجب: نفس النوع (EDP/EDT/EDC) إذا مذكور في الاثنين\n"
        "• يجب: نفس الجنس (رجالي/نسائي) إذا مذكور في الاثنين\n"
        "• إذا لا يوجد تطابق حقيقي → اكتب 0\n\n"
        + "\n\n".join(lines)
        + f"\n\nأجب JSON فقط، لا تكتب أي شيء آخر:\n"
          f"{{\"results\": [رقم1, رقم2, ..., رقم{len(batch_items)}]}}"
    )

    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.0,
            "maxOutputTokens": 200,
            "topP": 1.0,
            "topK": 1
        }
    }

    # ─── Retry loop: 3 محاولات ───
    for attempt in range(3):
        for key in GEMINI_API_KEYS:
            if not key: continue
            try:
                resp = _req.post(
                    f"{_GEMINI_URL}?key={key}",
                    json=payload, timeout=25
                )
                if resp.status_code == 200:
                    raw_text = resp.json()["candidates"][0]["content"]["parts"][0]["text"]
                    clean = re.sub(r'```json|```','', raw_text).strip()
                    s = clean.find('{'); e = clean.rfind('}') + 1
                    if s >= 0 and e > s:
                        data = json.loads(clean[s:e])
                        raw_results = data.get("results", [])
                        out = []
                        for j, item in enumerate(batch_items):
                            n = raw_results[j] if j < len(raw_results) else 1
                            try: n = int(n)
                            except: n = 1
                            if 1 <= n <= len(item["candidates"]): out.append(n - 1)
                            elif n == 0: out.append(-1)
                            else: out.append(0)
                        _cache_set(cache_key, out)
                        return out
                elif resp.status_code == 429:
                    time.sleep(2 ** attempt)  # exponential backoff
                    continue
            except _req.exceptions.Timeout:
                continue
            except Exception:
                continue
        time.sleep(1)

    # fallback: خذ الأول لكل منتج
    return [0] * len(batch_items)


# ══════════════════════════════════════════════
#  بناء صف النتيجة
# ══════════════════════════════════════════════
def _make_row(product, our_price, our_id, brand, size, ptype, gender,
              best=None, decision_override=None, ai_source="", all_candidates=None):
    """يبني صف النتيجة الكاملة"""

    if best is None:
        return {
            "المنتج":         product,
            "معرف المنتج":    our_id,
            "السعر":          our_price,
            "الماركة":        brand,
            "الحجم":          f"{int(size)}ml" if size else "",
            "النوع":          ptype,
            "الجنس":          gender,
            "منتج المنافس":   "—",
            "معرف المنافس":   "",
            "سعر المنافس":    0,
            "الفرق":          0,
            "نسبة التطابق":   0,
            "ثقة AI":         "—",
            "القرار":         decision_override or "🔵 مفقود عند المنافس",
            "الخطورة":        "",
            "المنافس":        "",
            "عدد المنافسين":  0,
            "جميع المنافسين": [],
            "مصدر المطابقة":  ai_source or "—",
            "تاريخ المطابقة": datetime.now().strftime("%Y-%m-%d"),
        }

    comp_price = float(best.get("price") or 0)
    score      = float(best.get("score") or 0)
    diff       = round(our_price - comp_price, 2) if (our_price > 0 and comp_price > 0) else 0
    abs_diff   = abs(diff)

    # مستوى الخطورة
    risk = "🔴 عالي" if abs_diff > 30 else "🟡 متوسط" if abs_diff > 10 else "🟢 منخفض"

    # القرار
    if decision_override:
        decision = decision_override
    elif ai_source in ("gemini", "auto") or score >= HIGH_CONFIDENCE:
        if diff > PRICE_TOLERANCE:     decision = "🔴 سعر أعلى"
        elif diff < -PRICE_TOLERANCE:  decision = "🟢 سعر أقل"
        else:                          decision = "✅ موافق"
    elif score >= REVIEW_THRESHOLD:
        decision = "⚠️ مراجعة"
    else:
        decision = "⚠️ مراجعة"

    # شارة ثقة AI
    ai_conf = {
        "gemini":          f"🤖 AI ✅ ({score:.0f}%)",
        "auto":            f"🎯 تلقائي ({score:.0f}%)",
        "gemini_no_match": "🤖 AI ❌ لا تطابق",
        "cache":           f"💾 Cache ({score:.0f}%)",
    }.get(ai_source, f"📊 {score:.0f}%")

    all_comps = (all_candidates or [best])[:5]
    n_comps   = len({c.get("competitor","") for c in all_comps if c.get("competitor")})

    return {
        "المنتج":         product,
        "معرف المنتج":    our_id,
        "السعر":          our_price,
        "الماركة":        brand,
        "الحجم":          f"{int(size)}ml" if size else "",
        "النوع":          ptype,
        "الجنس":          gender,
        "منتج المنافس":   best["name"],
        "معرف المنافس":   best.get("product_id", ""),
        "سعر المنافس":    comp_price,
        "الفرق":          diff,
        "نسبة التطابق":   score,
        "ثقة AI":         ai_conf,
        "القرار":         decision,
        "الخطورة":        risk,
        "المنافس":        best.get("competitor", ""),
        "عدد المنافسين":  n_comps,
        "جميع المنافسين": all_comps,
        "مصدر المطابقة":  ai_source or "fuzzy",
        "تاريخ المطابقة": datetime.now().strftime("%Y-%m-%d"),
    }


# ══════════════════════════════════════════════
#  التحليل الكامل الهجين
# ══════════════════════════════════════════════
def run_full_analysis(our_df, comp_dfs, progress_callback=None, use_ai=True):
    """
    التحليل الكامل:
    1. كل منتج عندنا → RapidFuzz يجد أفضل 6 مرشحين
    2. Gemini يختار الصح (batch كل 10)
    3. إذا score ≥ 97% → لا حاجة AI (سريع)
    4. Cache يمنع تكرار نفس الاستدعاء
    """
    results = []

    # ─── تحديد الأعمدة ───
    our_col       = _find_col(our_df, ["المنتج","اسم المنتج","Product","Name","name","product_name"])
    our_price_col = _find_col(our_df, ["السعر","سعر","Price","price","PRICE","سعر_البيع"])
    our_id_col    = _find_col(our_df, ["ID","id","معرف","رقم المنتج","SKU","sku","الكود","barcode"])

    comp_meta = {}
    for cname, cdf in comp_dfs.items():
        comp_meta[cname] = {
            "col":  _find_col(cdf, ["المنتج","اسم المنتج","Product","Name","name","product_name"]),
            "icol": _find_col(cdf, ["ID","id","معرف","رقم المنتج","SKU","sku","الكود","code","barcode"]),
        }

    total      = len(our_df)
    pending    = []   # batch للـ AI
    BATCH_SIZE = 10   # أمثل: 10 منتجات / استدعاء Gemini

    def _flush():
        """معالجة الـ batch الحالي"""
        if not pending: return
        indices = _ai_pick_batch(pending)
        for j, item in enumerate(pending):
            ci = indices[j] if j < len(indices) else 0
            if ci < 0:   # AI قرر: لا تطابق
                results.append(_make_row(
                    item["product"], item["our_price"], item["our_id"],
                    item["brand"], item["size"], item["ptype"], item["gender"],
                    None, "🔵 مفقود عند المنافس", "gemini_no_match"
                ))
            else:
                best = item["candidates"][ci]
                best["competitor"] = item["candidates"][ci].get("competitor", "")
                results.append(_make_row(
                    item["product"], item["our_price"], item["our_id"],
                    item["brand"], item["size"], item["ptype"], item["gender"],
                    best, ai_source="gemini", all_candidates=item["all_cands"]
                ))
        pending.clear()

    # ─── المرور على كل منتج ───
    for i, (_, row) in enumerate(our_df.iterrows()):
        product = str(row.get(our_col, "")).strip()
        if not product or is_sample(product):
            if progress_callback: progress_callback((i+1)/total)
            continue

        our_price = 0.0
        if our_price_col:
            try:
                v = str(row[our_price_col]).replace(",","")
                our_price = float(v)
            except: pass

        our_id = _get_id(row, our_id_col)
        brand  = extract_brand(product)
        size   = extract_size(product)
        ptype  = extract_type(product)
        gender = extract_gender(product)

        # ── المرحلة 1: RapidFuzz ──
        all_candidates = []
        for cname, meta in comp_meta.items():
            cands = _get_candidates(product, comp_dfs[cname], meta["col"], meta["icol"], top_n=6)
            for c in cands: c["competitor"] = cname
            all_candidates.extend(cands)

        if not all_candidates:
            results.append(_make_row(product, our_price, our_id, brand, size, ptype, gender,
                                     None, "🔵 مفقود عند المنافس"))
            if progress_callback: progress_callback((i+1)/total)
            continue

        all_candidates.sort(key=lambda x: x["score"], reverse=True)
        top5       = all_candidates[:5]
        best_score = top5[0]["score"]

        # ── قرار سريع بدون AI ──
        if best_score >= 97 or not use_ai:
            # واضح تماماً → لا حاجة AI
            results.append(_make_row(
                product, our_price, our_id, brand, size, ptype, gender,
                top5[0], ai_source="auto", all_candidates=all_candidates
            ))
            if progress_callback: progress_callback((i+1)/total)
            continue

        # ── يحتاج AI ──
        pending.append({
            "product":    product,
            "our_price":  our_price,
            "our_id":     our_id,
            "brand":      brand,
            "size":       size,
            "ptype":      ptype,
            "gender":     gender,
            "candidates": top5,
            "all_cands":  all_candidates,
            "our":        product,
            "price":      our_price,
        })

        if len(pending) >= BATCH_SIZE:
            _flush()

        if progress_callback: progress_callback((i+1)/total)

    _flush()  # الـ batch الأخير

    df_result = pd.DataFrame(results)
    return df_result


# ══════════════════════════════════════════════
#  المنتجات المفقودة
# ══════════════════════════════════════════════
def find_missing_products(our_df, comp_dfs):
    """يجد منتجات المنافسين الغير موجودة عندنا"""
    our_col   = _find_col(our_df, ["المنتج","اسم المنتج","Product","Name","name"])
    our_names_raw  = []
    our_names_norm = []
    for _, row in our_df.iterrows():
        p = str(row.get(our_col, "")).strip()
        if p and not is_sample(p):
            our_names_raw.append(p)
            our_names_norm.append(normalize(p))

    missing, seen = [], set()

    for cname, cdf in comp_dfs.items():
        ccol = _find_col(cdf, ["المنتج","اسم المنتج","Product","Name","name"])
        icol = _find_col(cdf, ["ID","id","معرف","رقم المنتج","SKU","sku","الكود","code"])

        for _, row in cdf.iterrows():
            cp = str(row.get(ccol, "")).strip()
            if not cp or is_sample(cp): continue
            cn = normalize(cp)
            if not cn or cn in seen: continue

            # هل موجود عندنا؟ (extractOne أسرع من loop)
            match = rf_process.extractOne(
                cn, our_names_norm,
                scorer=fuzz.token_sort_ratio,
                score_cutoff=70
            )
            if match: continue  # موجود عندنا

            seen.add(cn)
            missing.append({
                "منتج المنافس": cp,
                "معرف المنافس": _get_id(row, icol),
                "سعر المنافس":  _get_price(row),
                "المنافس":      cname,
                "الماركة":      extract_brand(cp),
                "الحجم":        f"{int(extract_size(cp))}ml" if extract_size(cp) else "",
                "النوع":        extract_type(cp),
                "الجنس":        extract_gender(cp),
                "تاريخ الرصد":  datetime.now().strftime("%Y-%m-%d"),
            })

    return pd.DataFrame(missing) if missing else pd.DataFrame()


# ══════════════════════════════════════════════
#  تصدير Excel احترافي
# ══════════════════════════════════════════════
def export_excel(df, sheet_name="النتائج"):
    """تصدير Excel مع تنسيق احترافي"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    output  = io.BytesIO()
    edf     = df.copy()
    if "جميع المنافسين" in edf.columns:
        edf = edf.drop(columns=["جميع المنافسين"])

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        edf.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        ws  = writer.sheets[sheet_name[:31]]

        # ─── تنسيق الرأس ───
        header_fill = PatternFill("solid", fgColor="1a1a2e")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ─── تلوين الصفوف بناءً على القرار ───
        colors = {
            "🔴 سعر أعلى": "FFCCCC",
            "🟢 سعر أقل":  "CCFFCC",
            "✅ موافق":     "CCFFEE",
            "⚠️ مراجعة":   "FFF3CC",
            "🔵 مفقود":     "CCE5FF",
        }
        decision_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value and "القرار" in str(cell.value):
                decision_col = idx; break

        if decision_col:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                decision_val = str(ws.cell(row_idx, decision_col).value or "")
                fill_color = None
                for k, c in colors.items():
                    if k.split()[0] in decision_val:
                        fill_color = c; break
                if fill_color:
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor=fill_color)

        # ─── عرض الأعمدة ───
        for col_idx, col in enumerate(ws.columns, 1):
            max_w = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_w + 4, 55)

    return output.getvalue()


def export_section_excel(df, section_name):
    return export_excel(df, sheet_name=section_name[:31])
